"""Excel catalog importer (Plan 1, Fase 2).

Replaces the Defontana ``sync_products`` pull: the product master is exported from
the ERP and imported here (manual upload or folder-watch). Acepta ``.xlsx``/``.xlsm``,
el ``.xls`` binario viejo y el HTML-disfrazado-de-``.xls`` del reporte "Informe de
Artículos" de Defontana (ver ``_defontana_records``). Rules:

- headers are matched by name (accent/case-insensitive synonyms); the **code (SKU)
  column is mandatory**;
- **all-or-nothing**: if any row is missing its SKU the whole file is rejected and
  nothing is written;
- otherwise every row is upserted by ``(tenant, sku)`` and its barcode added if new.

Also keeps a per-tenant last-import timestamp so a "no lo cargaste en 24 h" reminder
can fire (``catalog_staleness_check``).
"""
import hashlib
import io
import unicodedata
from datetime import timedelta, timezone
from html.parser import HTMLParser
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.core.logging import get_logger
from app.core.tenant_db import tenant_db
from app.core.utils import now_utc
from app.models import Collections
from app.models.notification import NotificationType
from app.models.product import BarcodeSource, BarcodeType
from app.services import notification_service

logger = get_logger(__name__)

# Recognized header synonyms (normalized: lowercased, accent-stripped, trimmed).
_HEADER_SYNONYMS: Dict[str, set] = {
    "sku": {"codigo", "code", "sku", "cod", "cod producto", "codigo producto", "cod. producto"},
    "name": {"nombre", "name", "descripcion", "detalle", "glosa", "producto"},
    "barcode": {"codigo de barras", "barcode", "ean", "ean13", "cod barra", "codigo barra",
                "cod de barras", "codigo de barra"},
    "category": {"categoria", "familia", "family", "rubro", "linea"},
    "unit": {"unidad", "unit", "um", "u.m.", "u. m."},
    "cost": {"costo", "cost", "costo neto"},
    "sale_price": {"precio", "precio venta", "price", "sale price", "precio de venta", "valor"},
}


def _norm(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return text.strip().lower()


def _num(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(" ", "")
    # Chilean formatting: "1.234,50" -> 1234.50
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _map_headers(row: Tuple) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, cell in enumerate(row):
        norm = _norm(cell)
        if not norm:
            continue
        for field, syns in _HEADER_SYNONYMS.items():
            if norm in syns and field not in mapping.values():
                mapping[idx] = field
                break
    return mapping


# OLE2 compound-file signature = formato viejo .xls (Excel 97-2003, BIFF).
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
# Pistas de HTML: el export "Informe de Artículos" de Defontana es HTML con
# extensión .xls (ni OLE2 ni ZIP), por eso openpyxl fallaba con "File is not a zip file".
_HTML_HINTS = (b"<html", b"<table", b"<!doctype html", b"<tr")


def _looks_like_html(data: bytes) -> bool:
    head = data[:4096].lstrip(b"\xef\xbb\xbf \t\r\n").lower()
    return head[:1] == b"<" and any(hint in head for hint in _HTML_HINTS)


class _HTMLTableRows(HTMLParser):
    """Extrae filas ``<tr>`` como tuplas de texto de celda, sin dependencias
    externas. colspan/rowspan se ignoran: se toman las celdas tal como vienen
    (el reporte de Defontana trae una fila plana de 8 columnas por producto)."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: List[Tuple] = []
        self._row: Optional[List[str]] = None
        self._cell: Optional[List[str]] = None

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).replace("\xa0", " ").strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(tuple(self._row))
            self._row = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def _rows_from_html(data: bytes) -> List[Tuple]:
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("cp1252", errors="replace")  # encoding típico de Defontana
    parser = _HTMLTableRows()
    parser.feed(text)
    return parser.rows


def _rows_from_bytes(data: bytes) -> List[Tuple]:
    """Extrae las filas de la primera hoja como tuplas, soportando .xlsx/.xlsm
    (openpyxl), el .xls binario viejo (xlrd) y el HTML-disfrazado-de-.xls que
    exporta Defontana (parser stdlib), detectados por los bytes de cabecera."""
    if data[:8] == _XLS_MAGIC:
        import xlrd  # solo se importa si llega un .xls binario

        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        return [tuple(sheet.row_values(r)) for r in range(sheet.nrows)]
    if _looks_like_html(data):
        return _rows_from_html(data)
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Layout específico: export "Informe de Artículos" de Defontana
# ---------------------------------------------------------------------------
# La columna 0 ("Artículo - Descripción") trae el código y el nombre pegados con
# guion y NO hay columna de código de barras / unidad / categoría / precio. El
# nombre es siempre multi-palabra y el código puede llevar guiones (ej. el prefijo
# 3M-70-2014-1). Se separa por el "primer segmento con espacio" (0 colisiones vs.
# el split ingenuo por primer guion) y se genera un EAN13 interno determinístico.


def _split_code_name(raw: Any) -> Tuple[str, Optional[str]]:
    text = " ".join(str(raw or "").split())
    if not text:
        return "", None
    segments = text.split("-")
    for i in range(1, len(segments)):
        if " " in segments[i].strip():  # 1er segmento multi-palabra => empieza el nombre
            return "-".join(segments[:i]).strip(), ("-".join(segments[i:]).strip() or None)
    if len(segments) > 1:  # hay guion pero ningún segmento multi-palabra
        return segments[0].strip(), ("-".join(segments[1:]).strip() or None)
    if " " in text:  # sin guion: separa por el primer espacio
        code, name = text.split(" ", 1)
        return code.strip(), (name.strip() or None)
    return text, None


def _ean13_check_digit(digits12: str) -> str:
    total = sum((3 if i % 2 else 1) * int(c) for i, c in enumerate(digits12))
    return str((10 - total % 10) % 10)


def _internal_ean13(sku: str) -> str:
    """EAN13 interno determinístico (prefijo 20 = uso interno, no colisiona con GTIN
    reales). Determinístico por SKU => re-importar NO duplica códigos de barras."""
    seq = int(hashlib.sha1(sku.encode("utf-8")).hexdigest(), 16) % (10 ** 10)
    base = "20" + str(seq).zfill(10)  # 12 dígitos
    return base + _ean13_check_digit(base)


def _is_defontana_header(cell: Any) -> bool:
    norm = _norm(cell)
    return norm.startswith("articulo") and "descrip" in norm


def _defontana_records(rows: List[Tuple]) -> Optional[List[Dict[str, Any]]]:
    """Si ``rows`` es el 'Informe de Artículos' de Defontana devuelve los productos
    (código + nombre + flag para generar barcode interno). Si no reconoce el
    layout devuelve ``None`` para que siga el importador genérico por cabeceras."""
    code_col = header_idx = None
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            if _is_defontana_header(cell):
                code_col, header_idx = c_idx, r_idx
                break
        if code_col is not None:
            break
    if code_col is None:
        return None

    records: List[Dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        if code_col >= len(row):
            continue
        raw = str(row[code_col] or "").strip()
        if not raw or _is_defontana_header(raw):
            continue
        if "-" not in raw and " " not in raw:
            continue  # línea suelta (subtotal/encabezado repetido), no es un producto
        code, name = _split_code_name(raw)
        if code:
            records.append({"sku": code, "name": name, "_gen_barcode": True})
    return records


def _parse_generic(rows: List[Tuple]) -> Tuple[bool, Dict[int, str], List[Tuple]]:
    header_map: Dict[int, str] = {}
    header_found = False
    data_rows: List[Tuple] = []
    for row in rows:
        if not header_found:
            candidate = _map_headers(row)
            if "sku" in candidate.values():
                header_map = candidate
                header_found = True
            continue
        data_rows.append(row)
    return header_found, header_map, data_rows


async def import_xlsx(tenant_id: str, data: bytes, actor: str, source: str = "upload") -> Dict[str, Any]:
    """Parse and upsert the catalog. Returns a report; never partially applies.

    Acepta tres formatos: .xlsx/.xlsm, .xls binario y el HTML-.xls de Defontana.
    El reporte de Defontana ("Informe de Artículos") se detecta por su columna
    "Artículo - Descripción" y se importa como catálogo (código + nombre + EAN13
    interno); NO trae stock, así que el stock no se toca."""
    try:
        rows = _rows_from_bytes(data)
    except Exception as exc:  # noqa: BLE001 - archivo corrupto / formato no soportado
        return {"applied": False, "rows": 0, "error": f"No se pudo leer el Excel: {exc}"}

    defontana = _defontana_records(rows)
    if defontana is not None:
        records = defontana
    else:
        header_found, header_map, data_rows = _parse_generic(rows)
        if not header_found:
            return {"applied": False, "rows": 0,
                    "error": "No se encontró la columna de código (SKU) en el Excel."}
        records = []
        rejected: List[Dict[str, Any]] = []
        for i, row in enumerate(data_rows, start=1):
            rec: Dict[str, Any] = {}
            for idx, field in header_map.items():
                rec[field] = row[idx] if idx < len(row) else None
            if not any(v not in (None, "") for v in rec.values()):
                continue  # fila vacía
            sku = str(rec.get("sku")).strip() if rec.get("sku") not in (None, "") else ""
            if not sku:
                rejected.append({"row": i, "reason": "Sin código (SKU)"})
                continue
            rec["sku"] = sku
            records.append(rec)
        if rejected:
            return {"applied": False, "rows": len(records) + len(rejected), "rejected": rejected,
                    "error": f"{len(rejected)} fila(s) sin código; no se aplicó ningún cambio."}

    if not records:
        return {"applied": False, "rows": 0, "error": "El Excel no tiene filas de productos."}

    db = tenant_db(tenant_id)
    now = now_utc()
    created = updated = barcodes_added = 0
    for rec in records:
        outcome, product_id = await _upsert_product(db, tenant_id, rec, actor, now)
        created += outcome == "created"
        updated += outcome == "updated"
        barcode = rec.get("barcode")
        if barcode not in (None, ""):
            if await _add_barcode(db, tenant_id, product_id, str(barcode).strip(), actor, now):
                barcodes_added += 1
        elif rec.get("_gen_barcode"):
            # Solo si el producto aún no tiene NINGÚN código de barras (idempotente).
            has_bc = await db[Collections.BARCODES].find_one(
                {"tenant_id": tenant_id, "product_id": product_id}, {"_id": 1}
            )
            if not has_bc and await _add_barcode(
                db, tenant_id, product_id, _internal_ean13(rec["sku"]), actor, now,
                bc_type=BarcodeType.INTERNAL.value, barcode_source=BarcodeSource.GENERATED.value,
            ):
                barcodes_added += 1

    report = {"applied": True, "rows": len(records), "created": created,
              "updated": updated, "barcodes_added": barcodes_added, "rejected": []}
    await _record_import(db, source, report)
    return report


async def _upsert_product(db, tenant_id: str, rec: Dict[str, Any], actor: str, now) -> Tuple[str, str]:
    sku = rec["sku"]
    existing = await db[Collections.PRODUCTS].find_one({"tenant_id": tenant_id, "sku": sku})
    name = str(rec["name"]).strip() if rec.get("name") not in (None, "") else (
        (existing or {}).get("name") or sku
    )
    doc: Dict[str, Any] = {
        "tenant_id": tenant_id,
        "sku": sku,
        "name": name,
        "category": (str(rec["category"]).strip() if rec.get("category") not in (None, "")
                     else (existing or {}).get("category") or "Sin categoría"),
        "unit": (str(rec["unit"]).strip() if rec.get("unit") not in (None, "")
                 else (existing or {}).get("unit") or "UN"),
        "is_active": True,
        "updated_at": now,
        "updated_by": actor,
    }
    if _num(rec.get("cost")) is not None:
        doc["cost"] = _num(rec.get("cost"))
    if _num(rec.get("sale_price")) is not None:
        doc["sale_price"] = _num(rec.get("sale_price"))

    if existing:
        await db[Collections.PRODUCTS].update_one({"_id": existing["_id"]}, {"$set": doc})
        return "updated", str(existing["_id"])

    doc.update({
        "erp_product_id": f"WMS-{sku}",
        "description": name,
        "brand": None,
        "uses_lots": False,
        "uses_serials": False,
        "is_service": False,
        "created_at": now,
        "created_by": actor,
    })
    result = await db[Collections.PRODUCTS].insert_one(doc)
    return "created", str(result.inserted_id)


async def _add_barcode(db, tenant_id: str, product_id: str, barcode: str, actor: str, now,
                       bc_type: Optional[str] = None,
                       barcode_source: str = BarcodeSource.MANUAL.value) -> bool:
    if not barcode:
        return False
    existing = await db[Collections.BARCODES].find_one({"tenant_id": tenant_id, "barcode": barcode})
    if existing:
        return False
    if bc_type is None:
        bc_type = (
            BarcodeType.EAN13.value if barcode.isdigit() and len(barcode) == 13
            else BarcodeType.SUPPLIER.value
        )
    await db[Collections.BARCODES].insert_one({
        "tenant_id": tenant_id,
        "product_id": product_id,
        "barcode": barcode,
        "type": bc_type,
        "source": barcode_source,
        "is_active": True,
        "created_at": now,
        "updated_at": now,
        "created_by": actor,
    })
    return True


# ---------------------------------------------------------------------------
# Staleness reminder ("no lo cargaste en 24 h")
# ---------------------------------------------------------------------------
def _aware(dt):
    if dt is None:
        return None
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


async def _record_import(db, source: str, report: Dict[str, Any]) -> None:
    await db[Collections.CATALOG_IMPORT_STATE].update_one(
        {},
        {"$set": {"last_import_at": now_utc(), "last_source": source, "last_report": report}},
        upsert=True,
    )


async def catalog_staleness_check(tenant_id: str, max_age_hours: int = 24) -> bool:
    """Emit a reminder if the catalog hasn't been imported within ``max_age_hours``.
    Deduped to at most one alert per window. Returns whether it alerted."""
    db = tenant_db(tenant_id)
    now = now_utc()
    state = await db[Collections.CATALOG_IMPORT_STATE].find_one({})
    last_import = _aware(state.get("last_import_at")) if state else None
    if last_import and (now - last_import) <= timedelta(hours=max_age_hours):
        return False
    last_alert = _aware(state.get("last_stale_alert_at")) if state else None
    if last_alert and (now - last_alert) < timedelta(hours=max_age_hours):
        return False
    await db[Collections.CATALOG_IMPORT_STATE].update_one(
        {}, {"$set": {"last_stale_alert_at": now}}, upsert=True
    )
    await notification_service.emit(
        tenant_id=tenant_id,
        notification_type=NotificationType.CATALOG_IMPORT.value,
        title="Catálogo sin actualizar",
        body=f"No se importó el catálogo en las últimas {max_age_hours} h.",
        entity_type="product",
    )
    return True
